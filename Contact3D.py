import numpy as np
import openpyxl as op

"""
make sheet without dead cells, arrest,
stop referencing openpyxl as much store as dtype 
"""

# import wb
infile = input("Enter .xlsx file: ")
savefile = input("What to save as: ")
# migrate_file = input("Enter the Migrate3D file: ")
contact_length = float(input("What distance between cells would you like to consider a contact? "))
arrested = float(input("What arrest coefficient value would you consider alive? "))

add_im = infile.find('.xlsx')
add_save = savefile.find('.xlsx')
# add_migrate = migrate_file.find('.xlsx')
if add_im == -1:
    importfile = infile + '.xlsx'
if add_save == -1:
    savefile = savefile + '.xlsx'
"""if add_migrate == -1:
    migrate_file = migrate_file + '.xlsx'"""

wb_in = op.load_workbook(infile)
sheet_in = wb_in.active
wb_out = op.Workbook()
sheet_out = wb_out.active
sheet_no_daughters = wb_out.create_sheet("Contacts without Mitosis")
sheet_no_dead = wb_out.create_sheet("Contacts with live cells")
# sheet_specific_interactions = wb_out.create_sheet("Contacts between", celltype1, celltype 2 )
# wb_migrate = op.load_workbook(migrate_file, read_only=True)
# migrate_sheet = wb_migrate['Summary']
sheet_out.cell(1, 1).value = "Time For Cell 1"
sheet_out.cell(1, 2).value = "Cell ID 1"
sheet_out.cell(1, 3).value = "Cell ID 2"
sheet_out.cell(1, 4).value = "Time for Cell 2"


# todo: add filter to see specific cell interactions (infected v not), will need to get different excel files add
#       them to a list then compare it there

def main():
    dict_ = create_dict()
    # arrest_dict = create_arrest_dict()
    detect_contacts(dict_)
    wb_out.save(savefile)


def create_dict():
    dict_of_cells = {}
    total = 1
    for row in range(2, sheet_in.max_row + 1):
        cell = sheet_in.cell(row, 2).value  # try to stop referencing sheet_in
        cell_chk = sheet_in.cell(row + 1, 2).value
        total += 1
        if cell != cell_chk:
            if cell not in dict_of_cells:
                time = sheet_in.cell(row, 3).value
                cell_start = row + 1 - (total - 1)
                dict_of_cells[cell] = [cell_start, row, time]
                # [0] is cell start, [1] is cell end
                total = 1
    return dict_of_cells


def create_arrest_dict():
    # create a dictionatry of only moving cells
    """arrst_dict = {}
    for row in range(2, migrate_sheet.max_row + 1):
        if migrate_sheet.cell(row, 22).value < arrested:
            arrst_dict[migrate_sheet.cell(row, 1).value] = [migrate_sheet.cell(row, 22).value]
        else:
            pass
    return arrst_dict"""


def detect_contacts(dict_):
    row = 2
    row_no_daughters = 2
    row_no_arrest = 2
    key_chk = list(dict_.keys())[0]
    for key_base in list(dict_.keys()):  # key to compare all to
        if key_chk != key_base:
            dict_.pop(key_chk)
            key_chk = key_base
        values = dict_[key_base]
        key_base_lowerbound = values[0]
        key_base_upperbound = values[1]
        for key_compare in dict_:  # keys that will be compared to key_base
            if key_compare == key_base:
                pass
            else:
                vals = dict_[key_compare]
                key_compare_lowerbound = vals[0]
                key_compare_upperbound = vals[1]
                # compare starting time point
                time_base_chk = vals[2]
                time_compare_chk = vals[2]
                # while loop to allow one to catch up to the other while making sure first instance
                while time_base_chk != time_compare_chk:
                    try:
                        time_base_chk = vals[2]
                        time_compare_chk = vals[2]
                        if time_base_chk > time_compare_chk:
                            key_compare_lowerbound += 1
                        elif time_compare_chk > time_base_chk:
                            key_base_lowerbound += 1
                        if time_base_chk == time_compare_chk:
                            break
                        if key_compare_lowerbound == key_compare_upperbound:
                            break
                        if key_base_lowerbound == key_base_upperbound:
                            break
                    except TypeError:
                        break

                for base_row, compare_row in zip(range(key_base_lowerbound, key_base_upperbound + 1),
                                                 range(key_compare_lowerbound, key_compare_upperbound + 1)):
                    time_base = sheet_in.cell(base_row, 3).value
                    time_compare = sheet_in.cell(compare_row, 3).value
                    try:
                        x_diff = np.abs(sheet_in.cell(base_row, 4).value - sheet_in.cell(compare_row, 4).value)
                        if x_diff <= contact_length:
                            y_diff = np.abs(
                                sheet_in.cell(base_row, 5).value - sheet_in.cell(compare_row, 5).value)
                            if y_diff <= contact_length:
                                z_diff = np.abs(sheet_in.cell(base_row, 6).value -
                                                sheet_in.cell(compare_row, 6).value)
                                if z_diff <= contact_length:  # all contacts final check
                                    sheet_out.cell(row, 1).value = time_base
                                    sheet_out.cell(row, 2).value = key_base
                                    sheet_out.cell(row, 3).value = key_compare
                                    sheet_out.cell(row, 4).value = time_compare
                                    row += 1
                                    if int(key_compare) - int(key_base) > 1:  # no daughters sheet
                                        sheet_no_daughters.cell(row_no_daughters, 1).value = time_base
                                        sheet_no_daughters.cell(row_no_daughters, 2).value = key_base
                                        sheet_no_daughters.cell(row_no_daughters, 3).value = key_compare
                                        sheet_no_daughters.cell(row_no_daughters, 4).value = time_compare
                                        row_no_daughters += 1

                                    """if key_base in arrest_dict and key_compare in arrest_dict and \
                                            np.abs(int(key_base) - int(key_compare)) > 1:
                                        # not daughter and alive only
                                        sheet_no_dead.cell(row_no_arrest, 1).value = time_base
                                        sheet_no_dead.cell(row_no_arrest, 2).value = key_base
                                        sheet_no_dead.cell(row_no_arrest, 3).value = key_compare
                                        sheet_no_dead.cell(row_no_arrest, 4).value = time_compare
                                        row_no_arrest += 1
                                    else:
                                        pass"""
                                else:
                                    pass
                            else:
                                pass
                        else:
                            pass
                    except ValueError:
                        pass

                    else:
                        pass

                key_base_lowerbound = values[0]
                key_base_upperbound = values[1]


main()
