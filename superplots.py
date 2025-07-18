import pandas as pd
import numpy as np
from scipy.stats import kruskal
from scikit_posthocs import posthoc_dunn
from plotly.subplots import make_subplots
import plotly.graph_objects as go
import plotly.colors
import glob
import re
import os

def superplots(input_pattern):
    # === Config ===
    input_files = sorted(glob.glob(input_pattern))

    base_name = os.path.basename(input_files[0])
    base_name = re.sub(r"-r\d+", "", base_name)
    html_output = base_name.replace("_Results.xlsx", "_Figures-Superplots.html")
    excel_output = base_name.replace("_Results.xlsx", "_Replicate-Stats.xlsx")

    # === Load and process data ===
    dfs = []
    replicate_sheets = []
    replicate_file_map = {}

    for file in input_files:
        try:
            df = pd.read_excel(file, sheet_name="Summary Features")
        except Exception as e:
            print(f"Skipping {file}: {e}")
            continue

        match = re.search(r"r(\d+)", os.path.basename(file), re.IGNORECASE)
        if match:
            rep_name = f"R{match.group(1)}"
        else:
            rep_name = os.path.splitext(os.path.basename(file))[0]

        df["Replicate"] = rep_name
        dfs.append(df)
        replicate_sheets.append(rep_name)
        replicate_file_map[rep_name] = file

    data = pd.concat(dfs, ignore_index=True)
    exclude_cols = {"Object ID", "Category", "Replicate"}
    features = [col for col in data.columns if col not in exclude_cols]

    stats_rows = []

    # === Prepare superplots ===
    ncols = 4
    nrows = int(np.ceil(len(features) / ncols))
    fig = make_subplots(
        rows=nrows,
        cols=ncols,
        subplot_titles=features,
        vertical_spacing=0.05,
        horizontal_spacing=0.05
    )

    replicate_colors = plotly.colors.qualitative.Plotly
    color_map = {rep: replicate_colors[i % len(replicate_colors)] for i, rep in enumerate(replicate_sheets)}

    first_plotted_feature = True

    for idx, feature in enumerate(features, 1):
        row = (idx - 1) // ncols + 1
        col = (idx - 1) % ncols + 1

        cat_reps = data.groupby(["Category", "Replicate"])[feature].apply(list).reset_index()

        medians = cat_reps.copy()
        medians["Median"] = medians[feature].apply(np.median)

        medians_df = (
            data.groupby(["Category", "Replicate"])[feature]
            .median()
            .reset_index()
        )

        for rep in replicate_sheets:
            rep_data = data[data["Replicate"] == rep]
            fig.add_trace(
                go.Box(
                    x=rep_data["Category"],
                    y=rep_data[feature],
                    name=rep,
                    boxpoints='all',
                    jitter=0.5,
                    pointpos=-1.8,
                    marker=dict(color=color_map[rep], opacity=0.5),
                    legendgroup=rep,
                    showlegend=first_plotted_feature,
                    boxmean=False,
                    line=dict(color=color_map[rep])
                ),
                row=row, col=col
            )

        overall_medians = data.groupby("Category")[feature].median().reset_index()
        fig.add_trace(
            go.Scatter(
                x=overall_medians["Category"],
                y=overall_medians[feature],
                mode="markers",
                marker_symbol="x",
                marker_color="black",
                marker_size=10,
                name="Overall Median",
                legendgroup="overall_median",
                showlegend=first_plotted_feature
            ),
            row=row, col=col
        )

        first_plotted_feature = False

        # === Stats ===
        low_variance = (medians_df[feature].var() == 0 or medians_df[feature].nunique() == 1)

        groups = [
            grp[feature].values
            for _, grp in medians_df.groupby("Category")
        ]
        n_kw = sum(len(g) for g in groups)

        if low_variance:
            stats_rows.append({
                "Feature": feature,
                "Test": "Kruskal-Wallis",
                "Comparison": "All Categories",
                "Statistic": "skipped",
                "p-value": "skipped (low variance)",
                "N": n_kw
            })
            for i in range(len(groups)):
                for j in range(i + 1, len(groups)):
                    cat_i = medians_df["Category"].unique()[i]
                    cat_j = medians_df["Category"].unique()[j]
                    n_i = len(groups[i])
                    n_j = len(groups[j])
                    stats_rows.append({
                        "Feature": feature,
                        "Test": "Holm-Bonferroni",
                        "Comparison": f"{cat_i} vs {cat_j}",
                        "Statistic": "skipped",
                        "p-value": "skipped (low variance)",
                        "N": n_i + n_j
                    })
        else:
            h, p_kw = kruskal(*groups)
            stats_rows.append({
                "Feature": feature,
                "Test": "Kruskal-Wallis",
                "Comparison": "All Categories",
                "Statistic": h,
                "p-value": p_kw,
                "N": n_kw
            })

            dunn_res = posthoc_dunn(medians_df, val_col=feature, group_col='Category', p_adjust='holm')
            for i in dunn_res.index:
                for j in dunn_res.columns:
                    if i < j:
                        n_i = medians_df[medians_df["Category"] == i][feature].count()
                        n_j = medians_df[medians_df["Category"] == j][feature].count()
                        stats_rows.append({
                            "Feature": feature,
                            "Test": "Holm-Bonferroni",
                            "Comparison": f"{i} vs {j}",
                            "Statistic": np.nan,
                            "p-value": dunn_res.loc[i, j],
                            "N": n_i + n_j
                        })

    # === Format and save figure ===
    fig.update_layout(
        height=int(500 * nrows),
        width=int(200 * len(replicate_sheets) * ncols),
        title_text="Superplots",
        title_x=0.5,
        title_font=dict(size=36),
        plot_bgcolor="white",
        paper_bgcolor="white",
        showlegend=True,
        legend=dict(
            orientation="v",
            x=1.02,
            y=1,
            xanchor="left",
            yanchor="top",
            bordercolor="black",
            borderwidth=1
        )
    )

    for i in range(1, nrows * ncols + 1):
        axis_x = f"xaxis{i}" if i > 1 else "xaxis"
        axis_y = f"yaxis{i}" if i > 1 else "yaxis"

        feature_idx = i - 1

        if feature_idx < len(features):
            feature = features[feature_idx]
            y_min = data[feature].min()
            y_max = data[feature].max()

            if y_min > 0:
                y_range = [0, y_max * 1.1]
            else:
                y_range = [y_min * 1.1, y_max * 1.1]
        else:
            y_range = None

        fig.layout[axis_x].update(
            showline=False,
            linecolor='black',
            showgrid=False,
            ticks='',
            tickcolor='black',
            ticklabelposition='outside',
            ticklabelstandoff=5
        )

        fig.layout[axis_y].update(
            showline=True,
            linecolor='black',
            linewidth=1,
            showgrid=False,
            ticks='outside',
            tickcolor='black',
            zeroline=True,
            zerolinecolor='black',
            zerolinewidth=1,
            range=y_range
        )

    fig.write_html(html_output)

    # === Save statistics ===
    stats_df = pd.DataFrame(stats_rows)
    stats_df.to_excel(excel_output, index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side

    wb = load_workbook(excel_output)
    ws = wb.active

    feature_col = 1
    prev_feature = None
    thick = Side(border_style="thick", color="000000")
    for row in range(2, ws.max_row + 1):
        feature = ws.cell(row=row, column=feature_col).value
        if prev_feature is not None and feature != prev_feature:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(top=thick)
        prev_feature = feature

    wb.save(excel_output)
