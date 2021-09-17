import openpyxl
import numpy as np
import pandas as pd
import statistics
"""
todo: to numeric 
make into a software where user can pick metrics they want, drop down?? 

"""
importfile = input("Name of the .XLSX file to import: ")
savefile = input("What to save as: ")
add_im = importfile.find('.xlsx')
add_save = savefile.find('.xlsx')
if add_im == -1:
    importfile = importfile + '.xlsx'
if add_save == -1:
    savefile = savefile + '.xlsx'
           
format_check = input("Do you wish to reformat the file? (y/n) ")
intervals = int(input("Time intervals to calculate (impacts run time severely): "))
arrest = float(input("Set arrest limit (assumes same units as XYZ coordinates): "))


if format_check == 'y':
    interpolate = input("Do you want to interpolate for missing data (y/n)? ")
    sort_1 = int(input("What column number would you like to sort by first? "))
    sort_2 = int(input("What column number would you like to sort by second? "))
    dem_2_chck = input("Replace Z dimension data with zeroes (i.e. convert 3D to 2D data)? (y/n) ")
    parent_id = int(input("What column number is your cell ID data in? "))
    time_col = int(input("What column number is your Time in? "))
    x_for = int(input("What column number is your X data in? "))
    y_for = int(input("What column number is your Y data in? "))
    if dem_2_chck == 'n':
        z_for = int(input("What column number is your Z data in? "))
    infile = pd.read_excel(importfile)
    df_mt_raw = pd.DataFrame(infile)
    df_mt_raw = df_mt_raw.sort_values(by=[df_mt_raw.columns[sort_1 - 1], df_mt_raw.columns[sort_2 - 1]],
                                      ascending=True)
    df_mt_raw.to_excel("formatted_" + importfile, index=False)

    importfile = "formatted_" + importfile
    wb = openpyxl.load_workbook(importfile)
    sheet = wb.active
    max_row = sheet.max_row
    mt_spot = False
    num_mt = 1
    # todo: need to figure out formatting / if statements make sure it catches all things
    for rows in range(2, max_row + 1):
        cells_id = sheet.cell(rows, 1).value
        cells_chck = sheet.cell(rows + 1, 1).value
        times = sheet.cell(rows, 2).value
        times_chk = sheet.cell(rows + 1, 2).value
        missing_data_cell_ID = []
        if cells_id == cells_chck:
            if times == times_chk:
                mt_spot = True
                num_mt += 1

            elif times != times_chk and mt_spot and cells_id == cells_chck:
                mt_spot = False
                x = []
                y = []
                z = []
                for i_s in range(num_mt):  # get all the values of a multitracked cell once spotted
                    x.append(sheet.cell(rows - i_s, 3).value)
                    y.append(sheet.cell(rows - i_s, 4).value)
                    z.append(sheet.cell(rows - i_s, 5).value)
                    sheet.cell(rows - i_s, 3).value = None
                    sheet.cell(rows - i_s, 4).value = None
                    sheet.cell(rows - i_s, 5).value = None

                sheet.cell(rows, 3).value = statistics.mean(x)
                sheet.cell(rows, 4).value = statistics.mean(y)
                sheet.cell(rows, 5).value = statistics.mean(z)
                num_mt = 1

        elif cells_id != cells_chck:
            total = 2
            if mt_spot:
                mt_spot = False
                x = []
                y = []
                z = []
                for i_s in range(num_mt):  # get all the values of a multitracked cell once spotted
                    x.append(sheet.cell(rows - i_s, 3).value)
                    y.append(sheet.cell(rows - i_s, 4).value)
                    z.append(sheet.cell(rows - i_s, 5).value)
                    sheet.cell(rows - i_s, 3).value = None
                    sheet.cell(rows - i_s, 4).value = None
                    sheet.cell(rows - i_s, 5).value = None

                sheet.cell(rows, 3).value = statistics.mean(x)
                sheet.cell(rows, 4).value = statistics.mean(y)
                sheet.cell(rows, 5).value = statistics.mean(z)
                num_mt = 1
            else:
                pass

    wb.save(importfile)
    infile = pd.read_excel(importfile)

    with pd.ExcelWriter(importfile, mode='a') as work_book:
        df = pd.DataFrame(infile)
        df = df.dropna()
        df.to_excel(work_book, sheet_name='Corrected', index=False)

    new_wrkbook = openpyxl.load_workbook(importfile)
    new_wrkbook.remove(new_wrkbook['Sheet1'])
    new_wrkbook.save(importfile)
    new_wrkbook.close()
    importex = pd.read_excel(importfile)
    df = pd.DataFrame(importex)
    df2 = pd.DataFrame()
    if dem_2_chck == 'n':
        df2['Parent ID'] = df[df.columns[parent_id-1]]
        df2["Time"] = df[df.columns[time_col-1]]
        df2["X Coordinate"] = df[df.columns[x_for-1]]
        df2["Y Coordinate"] = df[df.columns[y_for-1]]
        df2["Z Coordinate"] = df[df.columns[z_for-1]]
        df2.to_excel(importfile)
        print('Formatted file saved as: ', importfile)
    elif dem_2_chck == 'y':
        df2['Parent ID'] = df[df.columns[parent_id - 1]]
        df2["Time"] = df[df.columns[time_col - 1]]
        df2["X Coordinate"] = df[df.columns[x_for - 1]]
        df2["Y Coordinate"] = df[df.columns[y_for - 1]]
        df2["Z Coordinate"] = 0
        df2.to_excel(importfile)
        print('Formatted file saved as: ', importfile)

wb = openpyxl.load_workbook(importfile)
wb2 = openpyxl.Workbook()
sheet_in = wb.active
sheet = wb2.create_sheet('Calculations')
del wb2['Sheet']
sum_sheet = wb2.create_sheet("Summary")
ang_sheet = wb2.create_sheet("Angle Medians")
euclid_sheet = wb2.create_sheet("Euclidean Medians")


def main():
    sheet.cell(1, 8).value = "Cell ID"
    sheet.cell(1, 9).value = "Time"
    sheet.cell(1, 10).value = "Instantaneous Displacement"
    sheet.cell(1, 11).value = "Instantaneous Velocity"
    sheet.cell(1, 12).value = "Instantaneous Acceleration"
    sheet.cell(1, 13).value = "Instantaneous Velocity Filtered"
    sheet.cell(1, 14).value = "Instantaneous Acceleration Filtered"
    sheet.cell(1, 15).value = "Path length"
    sheet.cell(1, 16).value = "Total Displacement"
    sum_sheet.cell(1, 1).value = "Cell ID"
    sum_sheet.cell(1, 2).value = "Duration"
    sum_sheet.cell(1, 3).value = "Final Euclidean"
    sum_sheet.cell(1, 4).value = "Max Euclidean"
    sum_sheet.cell(1, 5).value = "Path Length"
    sum_sheet.cell(1, 6).value = "Straightness"
    sum_sheet.cell(1, 7).value = "TC Straightness"
    sum_sheet.cell(1, 8).value = "Displacement Ratio"
    sum_sheet.cell(1, 9).value = "Outreach Ratio"
    sum_sheet.cell(1, 10).value = "Velocity Mean"
    sum_sheet.cell(1, 11).value = "Velocity Median"
    sum_sheet.cell(1, 12).value = "Velocity SD"
    sum_sheet.cell(1, 13).value = "Velocity Mean (Filtered)"
    sum_sheet.cell(1, 14).value = "Velocity Median (Filtered)"
    sum_sheet.cell(1, 15).value = "Velocity SD (Filtered)"
    sum_sheet.cell(1, 16).value = "Acceleration Mean"
    sum_sheet.cell(1, 17).value = "Acceleration Median"
    sum_sheet.cell(1, 18).value = "Acceleration SD"
    sum_sheet.cell(1, 19).value = "Acceleration Mean (Filtered)"
    sum_sheet.cell(1, 20).value = "Acceleration Median (Filtered)"
    sum_sheet.cell(1, 21).value = "Acceleration SD (Filtered)"
    sum_sheet.cell(1, 22).value = "Arrest Coefficient"
    sum_sheet.cell(1, 23).value = "Angle LOBF Slope"
    sum_sheet.cell(1, 24).value = "Angle LOBF R2"
    sum_sheet.cell(1, 25).value = "Overall Angle Median"
    sum_sheet.cell(1, 26).value = "Euclidean LOBF Slope"
    sum_sheet.cell(1, 27).value = "Euclidean LOBF R2"
    sum_sheet.cell(1, 28).value = "Overall Euclidean Median"
    ang_sheet.cell(1, 1).value = 'Cell Id'
    euclid_sheet.cell(1, 1).value = 'Cell Id'
    col_add = 2

    # sets up variable euclid and angle input
    ang_col = intervals + 17
    for i in range(2, intervals + 1):
        sheet.cell(1, 15 + i).value = "Euclidean " + str(i) + " TP"
        if i % 2 == 1:
            sheet.cell(1, ang_col).value = "Angle " + str(i) + " TP"
            ang_col += 1

    # now set up angle filtered columns
    ang_col_calc = ang_col
    for i in range(2, intervals + 1):
        if i % 2 == 1:
            sheet.cell(1, ang_col_calc).value = "Angle Filtered " + str(i) + " TP"
            ang_col_calc += 1

    for t in range(2, intervals + 1):
        euclid_sheet.cell(1, t).value = "Euclidean Median " + str(t)
        if t % 2 == 0:
            pass
        else:
            ang_sheet.cell(1, col_add).value = "Median Angle " + str(t)
            col_add += 1

    total_ = 2
    cell_id1 = sheet_in.cell(1, 2).value
    max_row_in = sheet_in.max_row

    for row in range(2, max_row_in + 1):
        cell = sheet_in.cell(row, 2).value
        sheet.cell(row, 8).value = cell
        if cell_id1 == cell:
            time_points(row)
            displacement_and_pathlength(row, cell, cell_id1)
            total_disp(row, cell, total_)
            euclidian_in(row, total_)
            angle(total_, row, cell, cell_id1)
            cell_id1 = cell
            total_ = total_ + 1
        elif cell_id1 != cell:
            total_ = 2
            time_points(row)
            displacement_and_pathlength(row, cell, cell_id1)
            total_disp(row, cell, total_)
            euclidian_in(row, total_)
            angle(total_, row, cell, cell_id1)
            cell_id1 = cell
            total_ = total_ + 1

    wb2.save(savefile)
    total2 = 2
    cell_id2 = sheet.cell(2, 8).value
    for row in range(2, max_row_in + 1):
        cell2 = sheet.cell(row, 8).value
        if cell_id2 == cell2:
            velocity_and_acceleration_calculation(row, arrest, cell_id2, cell2)
            total2 += 1
            cell_id2 = cell2
        elif cell_id2 != cell2:
            velocity_and_acceleration_calculation(row, arrest, cell_id2, cell2)
            total2 = 2
            cell_id2 = cell2

    wb2.save(savefile)
    main_2(col_add)  # summary sheet main
    wb2.save(savefile)
    
    if format_check == 'y':
        print("Program completed: \n Formatted data saved as: ", importfile, 'Results saved as: ', savefile)
    else:
        print("Results saved as: ", savefile)


def time_points(row):
    time = sheet_in.cell(row, 3).value
    sheet.cell(row, 9).value = time


def displacement_and_pathlength(row, cell, cell_id1):
    if row < 2 or cell_id1 != cell:
        sheet.cell(row, 15).value = 0
        sheet.cell(row, 10).value = 0

    else:
        pl = sheet.cell(row - 1, 15).value
        cell_x_pos_0 = sheet_in.cell(row - 1, 4).value
        cell_y_pos_0 = sheet_in.cell(row - 1, 5).value
        cell_z_pos_0 = sheet_in.cell(row - 1, 6).value

        cell_x_pos = sheet_in.cell(row, 4).value
        cell_y_pos = sheet_in.cell(row, 5).value
        cell_z_pos = sheet_in.cell(row, 6).value

        if cell_x_pos is None or cell_z_pos is None:
            pass
        else:
            xd = (cell_x_pos - cell_x_pos_0) ** 2

            yd = (cell_y_pos - cell_y_pos_0) ** 2

            zd = (cell_z_pos - cell_z_pos_0) ** 2

            sum_coords = (xd + yd + zd)
            total_d = np.sqrt(sum_coords)

            # reset initial x for subtraction to prior

            path_length = pl + total_d

            sheet.cell(row, 10).value = total_d

            sheet.cell(row, 15).value = path_length


def total_disp(row, cell, total__):
    cell_chck = sheet_in.cell(row + 1, 2).value

    if total__ == 2:
        sheet.cell(row, 16).value = 0

    elif cell_chck != cell:
        newtotal = total__ - 2
        cell_x_pos_0 = sheet_in.cell(row - newtotal, 4).value
        cell_y_pos_0 = sheet_in.cell(row - newtotal, 5).value
        cell_z_pos_0 = sheet_in.cell(row - newtotal, 6).value
        for i in range(newtotal):
            cell_x_pos = sheet_in.cell(row - i, 4).value
            cell_y_pos = sheet_in.cell(row - i, 5).value
            cell_z_pos = sheet_in.cell(row - i, 6).value

            if cell_z_pos is None or cell_z_pos_0 is None:
                pass
            else:
                xtd = (cell_x_pos - cell_x_pos_0) ** 2

                ytd = (cell_y_pos - cell_y_pos_0) ** 2

                ztd = (cell_z_pos - cell_z_pos_0) ** 2

                sum_coords = (xtd + ytd + ztd)
                total_d = np.sqrt(sum_coords)
                sheet.cell(row - i, 16).value = total_d


def euclidian_in(row, total_euc):
    if total_euc < 4:
        pass
    elif total_euc <= intervals + 1:
        for i in range(total_euc - 3):
            euclidiancalc(17 + i, 2 + i, row)
    else:
        for i in range(intervals - 1):
            euclidiancalc(17 + i, 2 + i, row)


def euclidiancalc(colout_val, col_mod, row):
    try:
        x_0 = sheet_in.cell(row - col_mod, 4).value
        y_0 = sheet_in.cell(row - col_mod, 5).value
        z_0 = sheet_in.cell(row - col_mod, 6).value

        if type(x_0) is None or type(x_0) == str or z_0 is None:
            pass

        else:
            x_1 = sheet_in.cell(row, 4).value
            y_1 = sheet_in.cell(row, 5).value
            z_1 = sheet_in.cell(row, 6).value

            x_val = x_1 - x_0
            y_val = y_1 - y_0
            z_val = z_1 - z_0

            # calculate sqrd values
            x_sqrd = x_val ** 2
            y_sqrd = y_val ** 2
            z_sqrd = z_val ** 2

            euclidian_dist = np.sqrt(x_sqrd + y_sqrd + z_sqrd)

            sheet.cell(row, colout_val).value = euclidian_dist
    except ValueError:
        pass


def angle(total_ang, row, cell, cell_id1):
    if total_ang < 4 or cell_id1 != cell:
        pass
    else:
        col_a = intervals + 17
        s = 0
        calc_col = col_a + intervals
        arrst_multi = 1
        col_modif = 1
        even_mod = 0

        if intervals % 2 == 0:
            even_mod = 1

        for odd in range(intervals):
            if odd % 2 == 1:
                col_modif += 1

        for i in range(intervals - even_mod):
            euclid = 17 + i
            if i % 2 != 0:
                anglecalc(row, s + 1, col_a, calc_col - col_modif, euclid, arrest * arrst_multi, total_ang, i)
                col_a += 1
                calc_col += 1
                s += 1
                arrst_multi += 1
            else:
                pass


def anglecalc(row, sub, col_ang, col_calc, euclid_col, arrst_passed, total_ang, i):
    try:
        sub2 = sub * 2
        if i + 3 > total_ang:
            pass

        else:
            xt0 = sheet_in.cell(row, 4).value
            yt0 = sheet_in.cell(row, 5).value
            zt0 = sheet_in.cell(row, 6).value
            xt_1 = sheet_in.cell(row - sub, 4).value
            yt_1 = sheet_in.cell(row - sub, 5).value
            zt_1 = sheet_in.cell(row - sub, 6).value
            xt_2 = sheet_in.cell(row - sub2, 4).value
            yt_2 = sheet_in.cell(row - sub2, 5).value
            zt_2 = sheet_in.cell(row - sub2, 6).value
            # get the magnitude of each vector
            if xt_2 is None or type(xt_2) == str or zt0 is None:
                pass

            else:
                xmc = xt0 - xt_1
                ymc = yt0 - yt_1
                zmc = zt0 - zt_1
                xmp = xt_1 - xt_2
                ymp = yt_1 - yt_2
                zmp = zt_1 - zt_2
                vec_c = [xmc, ymc, zmc]
                vec_p = [xmp, ymp, zmp]
                # get the angle of the vectors
                vec_cnorm = vec_c / np.linalg.norm(vec_c)
                vec_pnorm = vec_p / np.linalg.norm(vec_p)
                angle_c = np.arccos(np.clip(np.dot(vec_cnorm, vec_pnorm), -1.0, 1.0))
                angle_deg = angle_c * 180 / np.pi
                sheet.cell(row, col_ang).value = angle_deg
                # now calc difference between angles by euclid
                euclid = sheet.cell(row, euclid_col).value
                euclid_prev = sheet.cell(row - sub, euclid_col).value
                if euclid is None or euclid_prev is None:
                    pass
                elif euclid > arrst_passed and euclid_prev > arrst_passed:
                    calc = np.absolute(angle_deg)
                    sheet.cell(row, col_calc).value = calc
    except ValueError:
        pass


def velocity_and_acceleration_calculation(row, arrst_passed, cell, cell_id1):
    if row <= 2 or cell != cell_id1:
        sheet.cell(row, 11).value = None
        sheet.cell(row, 13).value = None

    else:
        time_int = sheet.cell(row, 9).value - sheet.cell(row - 1, 9).value
        # instantaneous velocity and acceleration
        disp = sheet.cell(row, 10).value
        if disp is not None:
            sheet.cell(row, 11).value = disp / time_int
        accel_prior = sheet.cell(row - 1, 11).value
        if accel_prior is None:
            pass
        else:
            sheet.cell(row, 12).value = (sheet.cell(row, 11).value - accel_prior) / time_int
            disp_i = sheet.cell(row, 10).value
            if disp_i is None or disp_i <= arrst_passed:
                pass

            else:
                # filtered velocity
                sheet.cell(row, 13).value = sheet.cell(row, 10).value / time_int
                # acceleration filtered
                sheet.cell(row, 14).value = (sheet.cell(row, 11).value - sheet.cell(row - 1, 11).value) / time_int


def main_2(col_add):
    # main 2 contains all functions that will go to summary sheet
    total_2 = 2
    max_row_2 = sheet.max_row
    row_to_add = 2
    biggest_euclid = []
    time_under_lst = []
    path_lst = []
    accel_lst = []
    accel_filtered_lst = []
    time_under = []
    vel_lst = []
    vel_filtered_lst = []
    eight_chk = 0
    time_difference = sheet.cell(3, 9).value - sheet.cell(2, 9).value

    for row in range(2, max_row_2 + 1):
        cell_id1 = sheet.cell(row, 8).value
        cell_chk = sheet.cell(row + 1, 8).value
        if total_2 == 2:
            total_2 += 1
        elif cell_id1 == cell_chk:
            eight_count = sheet.cell(row, 13).value
            if eight_count is not None:
                eight_chk += 1
            max_and_final_and_id_and_bounds(row, row_to_add, biggest_euclid, cell_id1, cell_chk)
            path_len(row, row_to_add, path_lst)
            acceleration(row, accel_lst, accel_filtered_lst, cell_id1, cell_chk, row_to_add, eight_chk)
            velocity2(row, vel_lst, vel_filtered_lst, cell_id1, cell_chk, row_to_add, eight_chk)
            duration_and_under_arrest(row, total_2, time_under, row_to_add, cell_id1, cell_chk, time_under_lst,
                                      time_difference)
            total_2 += 1

        else:
            eight_count = sheet.cell(row, 13).value
            if eight_count is not None:
                eight_chk += 1
            max_and_final_and_id_and_bounds(row, row_to_add, biggest_euclid, cell_id1, cell_chk)
            mean_median_angle(row, total_2, row_to_add, col_add, intervals)
            euclid_lobf_and_euclidmedian(row, total_2, row_to_add, intervals)
            path_len(row, row_to_add, path_lst)
            acceleration(row, accel_lst, accel_filtered_lst, cell_id1, cell_chk, row_to_add,
                         eight_chk)
            duration_and_under_arrest(row, total_2, time_under, row_to_add, cell_id1, cell_chk, time_under_lst,
                                      time_difference)
            velocity2(row, vel_lst, vel_filtered_lst, cell_id1, cell_chk, row_to_add, eight_chk)
            # reset counters
            path_lst = []
            accel_lst = []
            accel_filtered_lst = []
            vel_lst = []
            vel_filtered_lst = []
            time_under_lst = []
            biggest_euclid = []
            time_under = []
            eight_chk = 0
            total_2 = 2
            row_to_add += 1

    wb.save(savefile)

    max_sum_row = sum_sheet.max_row

    for i in range(2, max_sum_row + 1):
        sum_stats(i)
        line_of_best_fit(i, col_add)
        median_of_medians()

    wb.save(savefile)


def max_and_final_and_id_and_bounds(row, row_to_add, biggest_euclid, cell_id1, cell_chk):
    current_euclid = sheet.cell(row, 16).value
    biggest_euclid.append(current_euclid)

    if cell_id1 != cell_chk:
        last_euclid = sheet.cell(row, 16).value
        sum_sheet.cell(row_to_add, 3).value = last_euclid
        cell_id = sheet.cell(row, 8).value
        sum_sheet.cell(row_to_add, 1).value = cell_id
        ang_sheet.cell(row_to_add, 1).value = cell_id
        euclid_sheet.cell(row_to_add, 1).value = cell_id
        max_euclid = max(biggest_euclid)
        sum_sheet.cell(row_to_add, 4).value = max_euclid


def path_len(row, row_to_add, path_lst):
    path = sheet.cell(row, 10).value
    if path is None:
        pass
    else:
        path_lst.append(path)
        sum_sheet.cell(row_to_add, 5).value = sum(path_lst)


def sum_stats(i):
    dur = sum_sheet.cell(i, 2).value
    final_euclid = sum_sheet.cell(i, 3).value
    max_euclid = sum_sheet.cell(i, 4).value
    pathl = sum_sheet.cell(i, 5).value
    if dur is None or final_euclid is None or pathl is None:
        pass
    else:
        sum_sheet.cell(i, 6).value = final_euclid / pathl
        sum_sheet.cell(i, 7).value = sum_sheet.cell(i, 6).value / (np.sqrt(dur))
        sum_sheet.cell(i, 8).value = final_euclid / max_euclid
        sum_sheet.cell(i, 9).value = max_euclid / pathl


def mean_median_angle(row, total, row_to_add, col_add, ang_amt):
    angle_sheet_start = 17 + ang_amt + col_add - 2
    col_in_sum = 2
    for col in range(angle_sheet_start, angle_sheet_start + col_add):
        lst = []
        for i in range(0, total):
            angle_add = sheet.cell(row - i, col).value
            if angle_add is None or angle_add == 0:
                pass
            else:
                lst.append(angle_add)
            if i == total - 2:
                if len(lst) >= 10:
                    medi = statistics.median(lst)
                    ang_sheet.cell(row_to_add, col_in_sum).value = medi
                    col_in_sum += 1
                else:
                    pass


def median_of_medians():
    max_med_sheet = ang_sheet.max_row
    max_col = ang_sheet.max_column
    for row_ in range(2, max_med_sheet + 1):
        medi_lst = []
        for col_ in range(2, max_col + 1):
            value = ang_sheet.cell(row_, col_).value
            if col_ == max_col:
                if medi_lst:
                    sum_sheet.cell(row_, 25).value = statistics.median(medi_lst)  # overall angle median
                else:
                    pass
            if value is not None:
                medi_lst.append(value)


def velocity2(row, vel_lst, vel_filter_lst, cell_id1, cell_chk, row_to_add, eight_chk):
    vel = sheet.cell(row, 11).value
    vel_filter = sheet.cell(row, 13).value
    if cell_id1 == cell_chk:
        if vel is None:
            pass
        else:
            vel_lst.append(vel)

        if vel_filter is None:
            pass
        else:
            vel_filter_lst.append(vel_filter)

    elif cell_id1 != cell_chk:
        try:
            sum_sheet.cell(row_to_add, 10).value = statistics.mean(vel_lst)
            sum_sheet.cell(row_to_add, 11).value = statistics.median(vel_lst)
            sum_sheet.cell(row_to_add, 12).value = statistics.stdev(vel_lst)

        except statistics.StatisticsError:
            pass

        if vel_filter_lst and eight_chk >= 8:
            try:
                sum_sheet.cell(row_to_add, 13).value = statistics.mean(vel_filter_lst)
                sum_sheet.cell(row_to_add, 14).value = statistics.median(vel_filter_lst)
                sum_sheet.cell(row_to_add, 15).value = statistics.stdev(vel_filter_lst)

            except statistics.StatisticsError:
                pass

    elif eight_chk <= 8:
        sum_sheet.cell(row_to_add, 10).value = statistics.mean(vel_lst)
        sum_sheet.cell(row_to_add, 11).value = statistics.median(vel_lst)
        sum_sheet.cell(row_to_add, 12).value = statistics.stdev(vel_lst)
        sum_sheet.cell(row_to_add, 13).value = None
        sum_sheet.cell(row_to_add, 14).value = None
        sum_sheet.cell(row_to_add, 15).value = None


def acceleration(row, accel_lst, accel_filtered_lst, cell_id1, cell_chk, row_to_add, eight_chk):
    accel = sheet.cell(row, 12).value
    accel_filter = sheet.cell(row, 14).value
    if accel is None:
        pass
    elif cell_id1 == cell_chk:
        accel_lst.append(accel)
        if accel_filter is None:
            pass
        else:
            accel_filtered_lst.append(accel_filter)
    elif cell_id1 != cell_chk:
        try:
            sum_sheet.cell(row_to_add, 16).value = statistics.mean(accel_lst)
            sum_sheet.cell(row_to_add, 17).value = statistics.median(accel_lst)
            sum_sheet.cell(row_to_add, 18).value = statistics.stdev(accel_lst)
            if accel_filtered_lst and eight_chk >= 8:
                sum_sheet.cell(row_to_add, 19).value = statistics.mean(accel_filtered_lst)
                sum_sheet.cell(row_to_add, 20).value = statistics.median(accel_filtered_lst)
                sum_sheet.cell(row_to_add, 21).value = statistics.stdev(accel_filtered_lst)

        except statistics.StatisticsError:
            pass

    elif eight_chk <= 8:
        sum_sheet.cell(row_to_add, 16).value = statistics.mean(accel_lst)
        sum_sheet.cell(row_to_add, 17).value = statistics.median(accel_lst)
        sum_sheet.cell(row_to_add, 18).value = statistics.stdev(accel_lst)
        sum_sheet.cell(row_to_add, 19).value = None
        sum_sheet.cell(row_to_add, 20).value = None
        sum_sheet.cell(row_to_add, 21).value = None


def duration_and_under_arrest(row, total, time_under, row_to_add, cell_id1, cell_chk, time_under_lst, time_difference):
    time_under_lst.append(sheet.cell(row, 9).value - sheet.cell(row - 1, 9).value)
    # time is the difference between two time points
    if cell_chk != cell_id1:
        dur_end = sum(time_under_lst)
        newtotal = total - 2
        dur_total = dur_end
        sum_sheet.cell(row_to_add, 2).value = dur_total
        for i in range(newtotal):  # calculates arrest coefficient in summary sheet
            disp = sheet.cell(row - i, 10).value
            if disp <= arrest and disp is not None:
                time_under.append(sheet.cell(row - i, 9).value - sheet.cell(row - (i+1), 9).value)

        if time_under:
            sum_sheet.cell(row_to_add, 22).value = (sum(time_under)) / dur_total

        else:
            sum_sheet.cell(row_to_add, 22).value = 0


def line_of_best_fit(i, col_add):
    y = []
    x = []
    calc_lobf_angles(i, x, y, col_add) # line of best fit for angles 


def calc_lobf_angles(row, x_lobf, y_lobf, col_add):
    for col in range(2, 2 + col_add):
        val = ang_sheet.cell(row, col).value
        if val is None:
            pass
        else:
            y_lobf.append(val)

    for i in range(1, len(y_lobf) + 1):
        x_lobf.append(i)

    x_lobf = np.array(x_lobf)
    y_lobf = np.array(y_lobf)

    if y_lobf.size >= 5: # todo: should make this variable
        model = np.polyfit(x_lobf, y_lobf, 1)
        sum_sheet.cell(row, 23).value = model[0]  # this is Angle LOBF
        cor_matrix = np.corrcoef(x_lobf, y_lobf)
        cor_xy = cor_matrix[0, 1]
        sum_sheet.cell(row, 24).value = cor_xy ** 2


def euclid_lobf_and_euclidmedian(row, total, row_to_add, ang_amt):
    euclid_lst = []
    euclid_lstm = []
    x = []
    col_euclid_add = 2

    # instant euclid
    for i in range(total - 3):
        euclid_lst.append(sheet.cell(row - i, 10).value)
    try:
        euclid_lstm.append(statistics.median(euclid_lst))
    except statistics.StatisticsError:
        pass

    for col in range(17, 17 + ang_amt):
        lst = []
        for z in range(total - 3):
            val = sheet.cell(row - z, col).value
            if val is None:
                pass
            else:
                lst.append(val)
            if z == total - 4:
                if len(lst) >= 10:
                    euclid_x_tp_median = statistics.median(lst)
                    euclid_sheet.cell(row_to_add, col_euclid_add).value = euclid_x_tp_median
                    euclid_lstm.append(euclid_x_tp_median)
                    col_euclid_add += 1
                else:
                    col_euclid_add += 1
    try:
        sum_sheet.cell(row_to_add, 28).value = statistics.median(euclid_lstm)  # Overall euclid median

    except statistics.StatisticsError:
        pass

    for i in range(1, len(euclid_lstm)+1):
        x.append(i)

    x = np.array(x)
    euclid_lstm = np.array(euclid_lstm)

    if len(euclid_lstm) >= 5:
        model = np.polyfit(x, euclid_lstm, 1)
        sum_sheet.cell(row_to_add, 26).value = model[0]  # euclidean lobf
        cor_matrix = np.corrcoef(x, euclid_lstm)
        cor_xy = cor_matrix[0, 1]
        sum_sheet.cell(row_to_add, 27).value = cor_xy ** 2

    else:
        pass


main()

